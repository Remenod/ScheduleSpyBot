import os
from urllib import response
import openpyxl
import requests
import subprocess
import databaseManager
from logger import log
from io import BytesIO
from botBase import bot
from dotenv import load_dotenv
from testModeVariable import TEST_MODE
from enumerations import Group, AdminPanel
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials


current_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(current_dir, '..', 'Secrets', 'KEYS.env')

load_dotenv(env_path)
if not TEST_MODE:
    SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")
else:
    SPREADSHEET_ID = os.getenv("SPREADSHEET_ID_T")

current_dir = os.path.dirname(os.path.abspath(__file__))
service_account_path = os.path.join(current_dir, "..", "Secrets", "schedulespybot-6e8cfdc17fcb.json")
SERVICE_ACCOUNT_FILE = os.path.abspath(service_account_path)

# Налаштування доступу до Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly', 'https://www.googleapis.com/auth/drive.readonly']
credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('sheets', 'v4', credentials=credentials)

def GetSheetGids() -> list:
    log("Loading gids...")
    try:
        sheet_metadata = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
                
        sheet_gids = []
        for sheet in sheet_metadata['sheets']:
            sheet_gids.append(sheet['properties']['sheetId'])
        
        return sheet_gids

    except Exception as e:
        print(f"Виникла помилка: {e}")
        return []
    finally:
        log("Gids loaded.")

gids = GetSheetGids()

def GetRightLines(value:str) -> str:
    if value is None:
        return None
    lines = str(value).split("\n")
    subjectLine = next((line for line in lines if line.strip()), None)
    auditoriumLine = next(
        (line for line in lines if line.strip().startswith("а.") or 
            line.strip().startswith("Наукова бібліотека")), None)
    if subjectLine is not None:
        if auditoriumLine is not None:
            return f"{subjectLine} $[]{auditoriumLine}"
        else:
            return f"{subjectLine}"
    else:
        return "помилка обробки розкладу"
    return value

def LoadWorkbook() -> openpyxl.workbook.workbook.Workbook:
    export_url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
    headers = {"Authorization": f"Bearer {credentials.token}"}
    response = requests.get(export_url, headers=headers)
    if response.status_code == 200:        
        return openpyxl.load_workbook(BytesIO(response.content))         
    else:
        text = f"Не вдалося завантажити файл. Код помилки: {response.status_code}"
        log(text)        
    
def GetSchedule(sheet_, columNum:Group, rawOutput:bool = False) -> str:
    sheet = sheet_
    output = f"{sheet.title}\n"
    row = 1
    while row <= sheet.max_row:
        cell = sheet[f"{columNum.value}{row}"]

        if cell.coordinate in sheet.merged_cells:        
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    main_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                    main_value = None
                    if(not rawOutput):
                        main_value = GetRightLines(main_cell.value)
                    else:
                        main_value = main_cell.value

                    if (row % 17 != 0 and row % 17 != 1) and main_value is not None:
                        output += f"{main_value}\n"
                    elif main_value is None:
                        output += f"Пара відсутня\n"
                    break
        else:                 
            cell_value = None
            if(not rawOutput):
                cell_value = GetRightLines(cell.value)
            else:
                cell_value = cell.value
            if (row % 17 != 0 and row % 17 != 1) and cell_value is not None:
                output += f"{cell_value}\n"
            elif cell_value is None:
                output += f"Пара відсутня\n"
        if row % 17 == 0:            
            row += 4
        elif row % 17 == 1:            
            row += 3
        else:
            row += 2
    return output

def CompareSchedules(input1:str, input2:str) -> str:
    current_dir = os.path.dirname(os.path.abspath(__file__))
    dll_path = os.path.join(current_dir, "..", "comparer", "bin", "Debug", "net8.0", "comparer.dll")

    dll_path = os.path.abspath(dll_path)

    try:
        run_result = subprocess.run(["dotnet", dll_path, input1, input2],
                                    capture_output=True,
                                    text=True,
                                    encoding="utf-8")
        if run_result.returncode != 0:
            log(f"Execution Error:{run_result.stderr.strip()}")
            return None        
        return run_result.stdout.strip()
    except FileNotFoundError:
        log("Error: Compiled executable not found.")
        raise Exception("Error: Compiled executable not found.")
        return None

def ParseComparerOutput(input:str) -> str:
    current_dir = os.path.dirname(os.path.abspath(__file__))
    dll_path = os.path.join(current_dir, "..", "parser", "bin", "Debug", "net8.0", "parser.dll")

    dll_path = os.path.abspath(dll_path)

    try:
        run_result = subprocess.run(["dotnet", dll_path, input],
                                    capture_output=True,
                                    text=True,
                                    encoding="utf-8")
        if run_result.returncode != 0:
            log(f"Execution Error:{run_result.stderr.strip()}")
            return None        
        return run_result.stdout.strip()
    except FileNotFoundError:
        log("Error: Compiled executable not found.")
        raise Exception("Error: Compiled executable not found.")
        return None

weekNums = databaseManager.GetAllSheetsNumbers()

def CompareAllGroups():
    log("Запускаю перевірку...")

    log("Завантажую таблицю з Google API...")
    workbook = LoadWorkbook()    

    weekNums = databaseManager.GetAllSheetsNumbers()
    
    if len(weekNums) == 0:
        log("Помилка. Не вдалося отримати записані тижні з бази данних")
        return None

    actualWeekNum = int(weekNums[0])
    lastWeekNum = int(workbook.worksheets[-1].title.split("т")[0])
    lastSavedWeekNum = int(weekNums[-1])

    sameAsOldWeekIndex = 1

    log("Перевіряю наявність нового тижня...")
    if(lastWeekNum != lastSavedWeekNum):
        try:
            currWeekNum = lastWeekNum            
            while currWeekNum != actualWeekNum:
                sameAsOldWeekIndex += 1
                currWeekNum = int(workbook.worksheets[-sameAsOldWeekIndex].title.split("т")[0])
                            
        except IndexError:
               log("Error: No old week found")
               return None

        finally:
            try:
                global gids
                res = bot.garanted_send_messages(databaseManager.GetAllUserIds(),"В розкладі з'явився новий тиждень.\n Для перегляду можете скористатися командою /schedule")
                if res != {}:
                    log(f"Помилки при повідомленні наявності нового тижня користувачам: \n{res}")
                log("В розкладі з'явились нові тижні")
                gids = GetSheetGids()
                temp = sameAsOldWeekIndex

                while (actualWeekNum+temp-1) != actualWeekNum:
                    for group in Group:
                        currSchedule = GetSchedule(workbook.worksheets[actualWeekNum+temp-2],group)
                        databaseManager.WriteSchedule(actualWeekNum+temp-1, group, f"{currSchedule}")

                    temp -= 1
            except Exception as e:
                log(e)
    else:
        log("нових тижнів не знайдено")
        sameAsOldWeekIndex = lastSavedWeekNum-actualWeekNum+1;


    log("Перевіряю зміни в кожній групі...")
    for group in Group:
        newSchedule = GetSchedule(workbook.worksheets[-sameAsOldWeekIndex], group)        
        oldSchedule = databaseManager.GetOldSchedule(actualWeekNum,group)

        comparerOut = CompareSchedules(oldSchedule, newSchedule)
        
        if comparerOut != "без змін":
            log(f"ВИЯВЛЕНІ ЗМІНИ В РОЗКЛАДІ ГРУПИ {group.name}")
            allCurrGroupUsers = databaseManager.GetAllUsersByGroup(group)
            allCurrGroupUsers.append(AdminPanel.groupId.value)
            if allCurrGroupUsers is not None or len(allCurrGroupUsers) != 0:                                        
                res1 = bot.garanted_send_messages(allCurrGroupUsers, f"*В розкладі виявлені зміни:*\n{ParseComparerOutput(comparerOut)}", "Markdown")
                bot.garanted_send_messages(allCurrGroupUsers, f"Для перегляду можете скористатися командою /schedule")

                log(f"Помилки при надсиланні змін користувачам: \n{res1}")

        databaseManager.WriteSchedule(actualWeekNum, group, newSchedule)
    log("Готово")