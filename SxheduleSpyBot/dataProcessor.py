import os
import subprocess
import requests
from dotenv import load_dotenv
from googleapiclient.discovery import build
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
from io import BytesIO
import enumerations as enums
from botBase import bot

load_dotenv(dotenv_path=r'../Secrets/KEYS.env')
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")
SERVICE_ACCOUNT_FILE = r'../Secrets/schedulespybot-86b7d86a2ebb.json'

# Налаштування доступу до Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly', 'https://www.googleapis.com/auth/drive.readonly']
credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('sheets', 'v4', credentials=credentials)   

def GetRightLines(value):
    if isinstance(value, str):
        lines = value.split("\n")
        subjectLine = next((line for line in lines if line.strip()), None)
        auditoriumLine = next(
            (line for line in lines if line.strip().startswith("а.") or 
             line.strip().startswith("Наукова бібліотека")), None)
        if subjectLine is not None:
            if auditoriumLine is not None:
                return f"{subjectLine} $[]{auditoriumLine}"
            else:
                return f"{subjectLine}"
        else:
            return "помилка обробки розкладу"
    return value

def LoadWorkbook():
    export_url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
    headers = {"Authorization": f"Bearer {credentials.token}"}
    response = requests.get(export_url, headers=headers)
    if response.status_code == 200:        
        return load_workbook(BytesIO(response.content))         
    else:
        raise Exception(f"Не вдалося завантажити файл. Код помилки: {response.status_code}")
    
def GetSchedule(sheet_, columNum = enums.Group.KC242_2.value, rawOutput = False):
    sheet = sheet_
    output = f"{sheet.title}\n"
    row = 1
    while row <= sheet.max_row:
        cell = sheet[f"{columNum}{row}"]

        if cell.coordinate in sheet.merged_cells:        
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    main_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                    main_value = None
                    if(not rawOutput):
                        main_value = GetRightLines(main_cell.value)
                    else:
                        main_value = main_cell.value
                    if (row % 17 != 0 and row % 17 != 1) and main_value is not None:
                        output += f"{main_value}\n"
                    elif main_value is None:
                        output += f"Пара відсутня\n"
                    break
        else:                 
            cell_value = None
            if(not rawOutput):
                cell_value = GetRightLines(cell.value)
            else:
                cell_value = cell.value
            if (row % 17 != 0 and row % 17 != 1) and cell_value is not None:
                output += f"{cell_value}\n"
            elif cell_value is None:
                output += f"Пара відсутня\n"
        if row % 17 == 0:            
            row += 4
        elif row % 17 == 1:            
            row += 3
        else:
            row += 2
    return output

def CompareSchedules(input1, input2):
    dll_path = os.path.join("../comparer", "bin", "Debug", "net8.0", "comparer.dll")       
    try:
        run_result = subprocess.run(["dotnet", dll_path, input1, input2],
                                    capture_output=True,
                                    text=True,
                                    encoding="utf-8")
        if run_result.returncode != 0:
            print("Execution Error:", run_result.stderr.strip())
            return None        
        return run_result.stdout.strip()

    except FileNotFoundError:
        print("Error: Compiled executable not found.")
        return None

def ParseComparerOutput(input):
    dll_path = os.path.join("../parser", "bin", "Debug", "net8.0", "parser.dll")       
    try:
        run_result = subprocess.run(["dotnet", dll_path, input],
                                    capture_output=True,
                                    text=True,
                                    encoding="utf-8")
        if run_result.returncode != 0:
            print("Execution Error:", run_result.stderr.strip())
            return None        
        return run_result.stdout.strip()
    except FileNotFoundError:
        print("Error: Compiled executable not found.")
        return None

def CompareAllGroups():    
    workbook = LoadWorkbook()
    oldSchedulesArr = []                                                       #get from db

    oldWeekName = oldSchedulesArr[0].split("\n")[0]
    sameAsOldWeekIndex = 1;

    if(workbook.worksheets[-1].title != oldWeekName):
        users = []                                                             #get from db
        for user in users:
            bot.send_message(user, "В розкладі з'явився новий тиждень!")
            
        try:
            while workbook.worksheets[-sameAsOldWeekIndex].title != oldWeekName:
                sameAsOldWeekIndex += 1
        except IndexError:
               print("Error: No old week found")
               return None

    groupIndex = 0
    for group in enums.Group:
        newSchedule = GetSchedule(workbook.worksheets[-sameAsOldWeekIndex], group.value)        
        comparerOut = CompareSchedules(newSchedule, oldSchedulesArr[groupIndex])
        groupIndex += 1
        if comparerOut != "без змін":
            allCurrGroupUsers = []                                               #get from db
            for user in allCurrGroupUsers:
                bot.send_message(user, f"{ParseComparerOutput(comparerOut)}")
            ###########                                                          #записати в базу в комірку розкладу групи {group.name} розклад {newSchedule}
