from calendar import weekday
import enum
# import telebot
from dotenv import load_dotenv
import os
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import requests
from io import BytesIO
from openpyxl import load_workbook

load_dotenv(dotenv_path=r'../Secrets/KEYS.env')
TELEGRAM_BOT_API = os.getenv("BOT_API")
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")
SERVICE_ACCOUNT_FILE = r'../Secrets/schedulespybot-86b7d86a2ebb.json'

# Налаштування доступу до Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly', 'https://www.googleapis.com/auth/drive.readonly']
credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('sheets', 'v4', credentials=credentials)

def DownloadSheet():
    export_url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
    headers = {"Authorization": f"Bearer {credentials.token}"}
    response = requests.get(export_url, headers=headers)
    if response.status_code == 200:
        return BytesIO(response.content)
    else:
        raise Exception(f"Не вдалося завантажити файл. Код помилки: {response.status_code}")

try:
    workbook = load_workbook(DownloadSheet())    
    print("Файл успішно завантажено")
except Exception as e:
    print(f"Помилка: {e}")

sheet = workbook.worksheets[-1]

class WeekDay(enum.Enum):   
    Понеділок = 0
    Вівторок  = 1
    Середа    = 2
    Четвер    = 3
    Пятниця   = 4
    Субота    = 5
    Неділя    = 6

def GetFirstNonEmptyLine(value):
    if isinstance(value, str):
        lines = value.split("\n")
        return next((line for line in lines if line.strip()), None)
    return value

row = 1
while row <= sheet.max_row:
    cell = sheet[f"F{row}"]

    if cell.coordinate in sheet.merged_cells:        
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                main_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                main_value = GetFirstNonEmptyLine(main_cell.value)
                if (row % 17 != 0 and row % 17 != 1) and main_value is not None:
                    print(f"{sheet[f'B{row}'].value:.0f} ПАРА: {main_value}")
                elif main_value is None:
                    print(f"{sheet[f'B{row}'].value:.0f} ПАРА: нема")
                break
    else:        
        cell_value = GetFirstNonEmptyLine(cell.value)
        if (row % 17 != 0 and row % 17 != 1) and cell_value is not None:
            print(f"{sheet[f'B{row}'].value:.0f} ПАРА: {cell_value}")
        elif cell_value is None:
            print(f"{sheet[f'B{row}'].value:.0f} ПАРА: нема")
    
    if row % 17 == 0:
        print(f"-----{WeekDay(row / 17).name}-----")
        row += 4
    elif row % 17 == 1:
        print(f"-----{WeekDay((row - 1) / 17).name}-----")
        row += 3
    else:
        row += 2



# def ColumRangeFormater(column):
#     column_name = []
#     while column > 0:
#         column -= 1
#         column_name.append(chr(65 + column % 26))  # 65 = 'A'
#         column //= 26
#     column_name.reverse()
#     return f"{''.join(column_name)}"



# bot = telebot.TeleBot(TELEGRAM_BOT_API)

# @bot.message_handler(commands=['start'])
# def start(message):
#     pass

# @bot.message_handler(func=lambda message: True)
# def send(message):
#     pass

# bot.polling(none_stop=True)
